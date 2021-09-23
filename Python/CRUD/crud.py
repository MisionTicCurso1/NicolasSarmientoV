
import openpyxl as ox
import datetime
import pandas as pd


ruta='C:\\Users\\Nicolas\\Documents\\Ejercicios MinTIC\\NicolasSarmientoVargas\\CRUD\\workbook.xlsx'
ruta=r'C:\\Users\\Nicolas\\Documents\\Ejercicios MinTIC\\NicolasSarmientoVargas\\CRUD\\workbook.xlsx'

pd.set_option('display.max_columns',99)
pd.set_option('display.width',6000)


def Agregar(ruta, data):

    excel_file=ox.load_workbook(ruta)
    data_sheet= excel_file['Report']
    data_sheet= data_sheet['A'+str(data_sheet.max_row+1):'F'+str(data_sheet.max_row+1)]
    hoja=excel_file.active
    
    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_final=6

    for i in data_sheet:
        if not(isinstance(i[0], int)):
            id=i[0].row
            hoja.cell(row=id, column=1).value=id-1
            hoja.cell(row=id, column=titulo).value= data['titulo']
            hoja.cell(row=id, column=descripcion).value=data['descripcion']
            hoja.cell(row=id, column=estado).value=data['estado']
            hoja.cell(row=id, column=fecha_inicio).value=data['fecha inicio']
            hoja.cell(row=id, column=fecha_final).value=data['fecha final']
        
        
        break

        

    excel_file.save(ruta)

    return


def leer(ruta, filtro, item=''):
    excel_file=ox.load_workbook(ruta)
    data_sheet= excel_file['Report']
    data_sheet= data_sheet['A2':'F'+str(data_sheet.max_row)]

    info={}


    if filtro=='todo':

        for i in data_sheet:
           
            if isinstance(i[0].value, int):
                info.setdefault(i[0].value, {'titulo':i[1].value, 'descripcion':i[2].value,'estado':i[3].value, 
                'fecha inicio':i[4], 'fecha final':i[5].value})
            
        print("****** REPORTE DE TODO ******")

    if filtro=='Id':

        for i in data_sheet:
            if i[0].value==int(item):
                info.setdefault(i[0].value, {'titulo':i[1].value, 'descripcion':i[2].value,'estado':i[3].value, 
                'fecha inicio':i[4], 'fecha final':i[5].value})

        print(f"****** REPORTE ID {item}  ******")

    if filtro=='titulo':
        
        for i in data_sheet:
            if i[1].value==item:
                info.setdefault(i[0].value, {'titulo':i[1].value, 'descripcion':i[2].value,'estado':i[3].value, 
                'fecha inicio':i[4], 'fecha final':i[5].value})

        print(f"****** REPORTE Titulo {item}  ******")

    if filtro=='estado':
        title='Estado'
        for i in data_sheet:
            if i[3].value==item:
                info.setdefault(i[0].value, {'titulo':i[1].value, 'descripcion':i[2].value,'estado':i[3].value, 
                'fecha inicio':i[4], 'fecha final':i[5].value})

        print(f"****** REPORTE Estados {item}  ******")

    if filtro=='f.inicio':
        
        for i in data_sheet:
            if i[4].value==item:
                info.setdefault(i[0].value, {'titulo':i[1].value, 'descripcion':i[2].value,'estado':i[3].value, 
                'fecha inicio':i[4], 'fecha final':i[5].value})

        print(f"****** REPORTE Fecha Inicio {item}  ******")

    if filtro=='f.final':
        
        for i in data_sheet:
            if i[5].value==item:
                info.setdefault(i[0].value, {'titulo':i[1].value, 'descripcion':i[2].value,'estado':i[3].value, 
                'fecha inicio':i[4], 'fecha final':i[5].value})

        print(f"****** REPORTE Fecha Fin {item}  ******")
    

    if info:

        DaTF=pd.DataFrame(info)
        DaTF.columns.name='ID'
        DaTF=DaTF.T
        print(DaTF)
    else:
        print(f'No se encontro {filtro} llamado {item}')
        
        
    
    
def Actualizar(ruta, id, dataup):
    excel_file=ox.load_workbook(ruta)
    data_sheet= excel_file['Report']
    data_sheet= data_sheet['A2':'F'+str(data_sheet.max_row+1)]
    hoja=excel_file.active
    found=False
    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_final=6

    for i in data_sheet:
        if i[0].value == int(id):
            found=True
            break
    
    if found:
        id=i[0].row
        if dataup['titulo']:
            hoja.cell(row=id, column=titulo).value= dataup['titulo']
        if dataup['descripcion']:
            hoja.cell(row=id, column=descripcion).value=dataup['descripcion']
        if dataup['estado']:
            hoja.cell(row=id, column=estado).value=dataup['estado']
        if dataup['fecha inicio']:
            hoja.cell(row=id, column=fecha_inicio).value=dataup['fecha inicio']
        if dataup['fecha final']:
            hoja.cell(row=id, column=fecha_final).value=dataup['fecha final']
    else:
        print('Id no encontrado')

    excel_file.save(ruta)


def Borrar(ruta, id):
    excel_file=ox.load_workbook(ruta)
    data_sheet= excel_file['Report']
    data_sheet= data_sheet['A2':'F'+str(data_sheet.max_row+1)]
    hoja=excel_file.active
    found=False
    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_final=6

    for i in data_sheet:
        if i[0].value == int(id):
            found=True
            break
    
    if found:
        id=i[0].row
        
        hoja.cell(row=id, column=1).value=''
        hoja.cell(row=id, column=titulo).value=''
        hoja.cell(row=id, column=descripcion).value=''
        hoja.cell(row=id, column=estado).value=''
        hoja.cell(row=id, column=fecha_inicio).value=''
        hoja.cell(row=id, column=fecha_final).value=''
    else:
        print('Id no encontrado')

    excel_file.save(ruta)


    

report={
    'titulo':'JSJSJS',
    'descripcion':'Se mimio en el trabajo',
    'estado':'Activo',
    'fecha inicio':'Hoy',
    'fecha final':'31/09/2021'
}






# while True:
#     print('****Menu CRUD***** \n')
#     print('Crear:      1')
#     print('Leer:       2')
#     print('Actualizar: 3')
#     print('Borrar:     4')
#     print('salir:      0 \n')

#     action=input('Ingrese accion a realizar: ')

#     if action=='1':
#         print('---------- CREAR----------')
#         registro_nuevo={
#             'titulo':'',
#             'descripcion':'',
#             'estado':'',
#             'fecha inicio':'',
#             'fecha final':''
#         }

#         print('Creando nuevo registro')
#         registro_nuevo['titulo']=input('Ingrese titulo: ')
#         registro_nuevo['descripcion']=input('Ingrese Descripcion: ')
#         registro_nuevo['estado']=input('Ingrese el estado: ')
#         registro_nuevo['fecha inicio']=input('Ingrese la fecha de inicio')
#         registro_nuevo['fecha final']=input('Ingrese la fecha final del reporte: ')

#         Agregar(ruta, registro_nuevo)
#         print('Haz creado un nuevo registro :)')

#     elif action=='2':
#         print('---------- LEER----------')
#         print('consultar todo:         0')
#         print('consultar Id:           1')
#         print('consultar titulo:       2')
#         print('consultar estado:       4')
#         print('consultar fecha inicio: 5')
#         print('consultar fecha final:  6')

#         consulta=input('Elija opcion a consultar')
#         if not (consulta=='0'):
#             option=input('Elija item a consultar:')

#         if consulta=='0':
#             leer(ruta, 'todo')
#         elif consulta=='1':
#             leer(ruta,'Id', option)
#         elif consulta=='2':
#             leer(ruta, 'titulo', option)
#         elif consulta=='3':
#             leer(ruta, 'estado', option)
#         elif consulta=='4':
#             leer(ruta, 'f.inicio', option)
#         elif consulta=='5':
#             leer(ruta, 'f.final', option)
#         else:
#             print('comando no valido')


        
#     elif action=='3':
#         print('---------- ACTUALIZAR----------')

        
#     elif action=='4':
#         print('---------- BORRAR----------')
        
#     elif action=='0':
#         print('hasta la proxima')
#         break
#     else:
#         print('Comando invalido... Vuelve a intentarlo')