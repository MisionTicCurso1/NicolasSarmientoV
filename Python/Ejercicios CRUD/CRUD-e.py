from os import read
from openpyxl import load_workbook
import datetime
import pandas as pd

pd.set_option('display.max_columns',99)
pd.set_option('display.width',9000)

ruta='C:\\Users\\Nicolas\\Documents\\Ejercicios MinTIC\\NicolasSarmientoVargas\\Ejercicios CRUD\\Covid-workbook.xlsx'
ruta=r'C:\\Users\\Nicolas\\Documents\\Ejercicios MinTIC\\NicolasSarmientoVargas\\Ejercicios CRUD\\Covid-workbook.xlsx'


def Add(ruta, new_info):
    excel_file=load_workbook(ruta)
    data_sheet= excel_file['Pacientes']
    data_sheet= data_sheet['A'+str(data_sheet.max_row+1):'J'+str(data_sheet.max_row+1)]
    hoja=excel_file.active

    nombre=2
    apellido=3
    gs=4
    rh=5
    edad=6
    vacunado=7
    sCovid=8
    creacion=9
    habilitado=10

    t=datetime.datetime.now()
    time=f'{t.year}-{t.month}-{t.day} {t.hour}:{t.minute}'

    for i in data_sheet:
        if not(isinstance(i[0], int)):
            id=i[0].row
            hoja.cell(row=id, column=1).value=id-1
            hoja.cell(row=id, column=nombre).value= new_info['nombre']
            hoja.cell(row=id, column=apellido).value=new_info['apellido']
            hoja.cell(row=id, column=gs).value=new_info['g.sanguineo']
            hoja.cell(row=id, column=rh).value=new_info['RH']
            hoja.cell(row=id, column=edad).value=int(new_info['edad'])
            hoja.cell(row=id, column=vacunado).value=new_info['vacunado']
            hoja.cell(row=id, column=sCovid).value=new_info['sCovid']
            hoja.cell(row=id, column=creacion).value=time
            hoja.cell(row=id, column=habilitado).value='si'
        
        break

    excel_file.save(ruta)

def Read(ruta, filtro, item1='', item2=''):
    
    excel_file=load_workbook(ruta)
    data_sheet= excel_file['Pacientes']
    data_sheet= data_sheet['A2':'J'+str(data_sheet.max_row)]

    reading={}


    if filtro=='todo':
        for i in data_sheet:
            if isinstance(i[0].value, int) and i[9].value=='si':
                reading.setdefault(i[0].value,{'nombre':i[1].value, 'apellidos':i[2].value,'g.sanguineo':i[3].value,
                'RH':i[4].value, 'Edad':i[5].value, 'vacunado':i[6].value, 'contagiado Covid':i[7].value, 'fecha creacion':i[8].value})
    elif filtro=='id':
        for i in data_sheet:
            if i[0].value==item1 and i[9].value=='si':
                reading.setdefault(i[0].value,{'nombre':i[1].value, 'apellidos':i[2].value,'g.sanguineo':i[3].value,
                'RH':i[4].value, 'Edad':i[5].value, 'vacunado':i[6].value, 'contagiado Covid':i[7].value,'fecha creacion':i[8].value})
    elif filtro=='nombre':
        for i in data_sheet:
            if i[1].value==item1 and i[2].value==item2 and i[9].value=='si':
                reading.setdefault(i[0].value,{'nombre':i[1].value, 'apellidos':i[2].value,'g.sanguineo':i[3].value,
                'RH':i[4].value, 'Edad':i[5].value, 'vacunado':i[6].value, 'contagiado Covid':i[7].value,'fecha creacion':i[8].value})        
    elif filtro=='RH':
        for i in data_sheet:
             if i[3].value==item1 and i[4].value==item2 and i[9].value=='si':
                reading.setdefault(i[0].value,{'nombre':i[1].value, 'apellidos':i[2].value,'g.sanguineo':i[3].value,
                'RH':i[4].value, 'Edad':i[5].value, 'vacunado':i[6].value, 'contagiado Covid':i[7].value,'fecha creacion':i[8].value})
    elif filtro=='edad':
        for i in data_sheet:
            
            if  i[9].value=='si' and int(i[5].value)>=int(item1) and int(i[5].value)<=int(item2):
                reading.setdefault(i[0].value,{'nombre':i[1].value, 'apellidos':i[2].value,'g.sanguineo':i[3].value,
                'RH':i[4].value, 'Edad':i[5].value, 'vacunado':i[6].value, 'contagiado Covid':i[7].value,'fecha creacion':i[8].value})
    elif filtro=='vacunado':
        for i in data_sheet:
             if i[6].value==item1 and i[9].value=='si':
                reading.setdefault(i[0].value,{'nombre':i[1].value, 'apellidos':i[2].value,'g.sanguineo':i[3].value,
                'RH':i[4].value, 'Edad':i[5].value, 'vacunado':i[6].value, 'contagiado Covid':i[7].value,'fecha creacion':i[8].value})
    elif filtro=='SCovid':
    
        for i in data_sheet:
             if i[7].value==item1 and i[9].value=='si':
                reading.setdefault(i[0].value,{'nombre':i[1].value, 'apellidos':i[2].value,'g.sanguineo':i[3].value,
                'RH':i[4].value, 'Edad':i[5].value, 'vacunado':i[6].value, 'contagiado Covid':i[7].value,'fecha creacion':i[8].value})

    if reading:

        DaTF=pd.DataFrame(reading)
        DaTF.columns.name='ID'
        DaTF=DaTF.T
        print(DaTF)
    else:
        print('No se encontro ningun objeto')


def Update(ruta, id, info_update):
    
    excel_file=load_workbook(ruta)
    data_sheet= excel_file['Pacientes']
    data_sheet= data_sheet['A2':'K'+str(data_sheet.max_row)]
    hoja=excel_file.active

    nombre=2
    apellido=3
    gs=4
    rh=5
    edad=6
    vacunado=7
    sCovid=8
    found=False

    t=datetime.datetime.now()
    time=f'{t.year}-{t.month}-{t.day} {t.hour}:{t.minute}'
    
    for i in data_sheet:
        if i[0].value == int(id) and i[9].value=='si':
            found=True
            break
    
    if found and i[9].value=='si':
        id=i[0].row
        if info_update['nombre']:
           hoja.cell(row=id, column=nombre).value= info_update['nombre']
        if info_update['apellido']:
            hoja.cell(row=id, column=apellido).value=info_update['apellido']
        if info_update['g.sanguineo']:
            hoja.cell(row=id, column=gs).value=info_update['g.sanguineo']
        if info_update['RH']:
            hoja.cell(row=id, column=rh).value=info_update['RH']
        if info_update['edad']!=0:
            hoja.cell(row=id, column=edad).value=info_update['edad']
        if info_update['vacunado']:
            hoja.cell(row=id, column=vacunado).value=info_update['vacunado']
        if info_update['sCovid']:
            hoja.cell(row=id, column=sCovid).value=info_update['sCovid']

        hoja.cell(row=id, column=11).value=time
    else:
        print('Id no encontrado')

    excel_file.save(ruta)
    
def Delete(ruta, id:int):
    excel_file=load_workbook(ruta)
    data_sheet= excel_file['Pacientes']
    data_sheet= data_sheet['A2':'J'+str(data_sheet.max_row)]
    hoja=excel_file.active

    found=False


    for i in data_sheet:
        if i[0].value == id and i[9].value=='si':
            found=True
            break

    if found:
        Id=i[0].row
        hoja.cell(row=Id, column=10).value='No'
    else:
        print('Id No encontrado')

    excel_file.save(ruta)



# paciente={
#     'nombre':'',
#     'apellido':'',
#     'g.sanguineo':'',
#     'RH':'-',
#     'edad':0,
#     'vacunado':'no',
#     'sCovid':'si'

# }


while True:
    print('Menu................................\n')
    print('Crear............................  1')
    print('Leer.............................  2')
    print('Editar............................ 3')
    print('Borrar............................ 4')
    print('Salir............................  0')

    option=int(input('Ingrese tarea a realizar: '))

    if option==1:
        print('...........................CREAR..........................')
        paciente_nuevo={
            'nombre':'',
            'apellido':'',
            'g.sanguineo':'',
            'RH':'',
            'edad':0,
            'vacunado':'',
            'sCovid':''
        }
        print('Para crear su nuevo registro:')
        paciente_nuevo['nombre']=input('Ingrese el Nombre del paciente: ')
        paciente_nuevo['apellido']=input('Ingrese el Apellido del paciente: ')
        paciente_nuevo['edad']=int(input('Ingrese la edad del paciente: '))
        
        print('Grupos sanguineos: \n1:A \n2:AB \n3:B \n4:O')
        element=int(input('Elija grupo sanguineo: '))
        if element==1:
            paciente_nuevo['g.sanguineo']='A'
        elif element==2:
            paciente_nuevo['g.sanguineo']='AB'
        elif element==3:
            paciente_nuevo['g.sanguineo']='B'
        elif element==4:
            paciente_nuevo['g.sanguineo']='O'

        print('RH: \n1:- \n2:+')
        element=int(input('Elija RH: '))

        if element==1:
            paciente_nuevo['RH']='-'
        elif element==2:
            paciente_nuevo['RH']='+'

        print('Esta vacunado: \n1: Si \n2: No ')
        element=int(input('Elija Estado de vacunacion: '))

        if element==1:
            paciente_nuevo['vacunado']='si'
        elif element==2:
            paciente_nuevo['vacunado']='no'

        element=int(input('Ha sufrido de Covid: \n1: Si \n2: No \n'))

        if element==1:
            paciente_nuevo['sCovid']='si'
        elif element==2:
            paciente_nuevo['sCovid']='no'

        Add(ruta,paciente_nuevo)
    

    elif option==2:
        print('............................LEER...........................')
        print('Opciones para consultar:\n1:TODO \n2:por ID \n3:Por Nombre')
        print('4:Por RH \n5:Por Rango de edades \n6:Por Estado de vacunacion \n7:Si padecio la Covid 19')

        element=int(input('Elija filtro a usar: '))

        if element==1:
            Read(ruta, 'todo')
        elif element==2:
            id=int('Id del paciente a consultar: ')
            Read(ruta, 'id',id)
        elif element==3:
            nombre=input('Nombre del paciente a consultar: ')
            apellido=input('Apellido del paciente: ')
            Read(ruta,'nombre',nombre, apellido)
        elif element==4:
            print('Grupos sanguineos: \n1:A \n2:AB \n3:B \n4:O')
            cod=int(input('Elija grupo sanguineo: '))
            if cod==1:
               GS='A'
            elif cod==2:
                GS='AB'
            elif cod==3:
                GS='B'
            elif cod==4:
                GS='O'
            
            print('RH: \n1: - \n2:+ ')
            cod=int(input('Elija RH: '))

            if cod==1:
                Rh='-'
            elif cod==2:
                Rh='+'
            
            Read(ruta,'RH',GS,Rh)
        elif element==5:
            Edad1=int(input('Ingrese la edad minima: '))
            Edad2=int(input('Ingrese la edad maxima: '))

            Read(ruta, 'edad', Edad1, Edad2)
        elif element==6:
            cod=int(input('Esta Vacunado: \n1: Si \n2: No '))

            if cod==1:
                vacunado='si'
            elif cod==2:
                vacunado='no'

            Read(ruta,'vacunado',vacunado)
        elif element==7:
            cod=int(input('Ha sufrido de Covid: \n1: Si \n2: No '))

            if cod==1:
                covid='si'
            elif cod==2:
                covid='no'

            Read(ruta,'SCovid',covid)
    elif option==3:
        print('...........................EDITAR..........................')
        cod=int(input('Desea leer antes de editar \n1:Si \n0:No '))
        if cod==1:
            Read(ruta,'todo')

        paciente_editado={
            'nombre':'',
            'apellido':'',
            'g.sanguineo':'',
            'RH':'',
            'edad':0,
            'vacunado':'',
            'sCovid':''
        }

        id=int(input('Elija el Id a Editar: '))
        nombre=int(input('Desea editar el nombre: \n1:si \n0:no '))
        Apellido=int(input('Desea editar el Apellido: \n1:si \n0:no '))
        Edad=int(input('Desea editar el Edad: \n1:si \n0:no'))
        Gsan=int(input('Desea editar el grupo sanguineo: \n1:si \n0:no' ))
        Rh=int(input('Desea editar el RH: \n1:si \n0:no '))
        vacunado=int(input('Desea editar el estado de vacunacion: \n1:si \n0:no '))
        sCovid=int(input('Desea editar si al paciente le dio Covid: \n1:si \n0:no '))

        if nombre==1:
            paciente_editado['nombre']=input('Ingrese el Nombre del paciente: ')
        if Apellido==1:
            paciente_editado['apellido']=input('Ingrese el Apellido del paciente: ')
        if Edad==1:
            paciente_editado['edad']=int(input('Ingrese la edad del paciente: '))
        if Gsan==1:
            print('Grupos sanguineos: \n1:A \n2:AB \n3:B \n4:O ')
            element=int(input('Elija grupo sanguineo: '))
            if element==1:
                paciente_editado['g.sanguineo']='A'
            elif element==2:
                paciente_editado['g.sanguineo']='AB'
            elif element==3:
                paciente_editado['g.sanguineo']='B'
            elif element==4:
                paciente_editado['g.sanguineo']='O'
        if Rh==1:
            print('RH: \n1:- \n2:+ \n')
            element=int(input('Elija RH: '))

            if element==1:
                paciente_editado['RH']='-'
            elif element==2:
                paciente_editado['RH']='+'
        if vacunado==1:
            print('Esta vacunado: \n1: Si \n2: No ')
            element=int(input('Elija Estado de vacunacion: '))

            if element==1:
                paciente_editado['vacunado']='si'
            elif element==2:
                paciente_editado['vacunado']='no'

        if sCovid==1:
            element=int(input('Ha sufrido de Covid: \n1: Si \n2: No \n'))

            if element==1:
                paciente_editado['sCovid']='si'
            elif element==2:
                paciente_editado['sCovid']='no'

        Update(ruta, id, paciente_editado)
        


    elif option==4:
        print('...........................BORRAR..........................')
        cod=int(input('Desea leer antes de Borrar \n1:Si \n0:No '))
        if cod==1:
            Read(ruta,'todo')

        id=int(input('Elija el Id a Borrar: '))

        Delete(ruta, id)


    elif option==0:
        option=int(input('desea salir?   1:si   0:No '))
        if option==1:
            print('hasta la proxima :)')
            break
        
    else:
        print('comando invalido')